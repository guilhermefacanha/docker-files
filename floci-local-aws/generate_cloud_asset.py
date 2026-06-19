#!/usr/bin/env python3
"""
generate_cloud_asset.py — Generate a DSF / Floci cloud-asset import spreadsheet.

Reads environment variables (with .env fallback, same priority as 00-env.sh) to
populate ARNs, account IDs, credentials and endpoints.  Outputs XLSX when
openpyxl is available, otherwise plain CSV with a note to convert manually.

Quick start
-----------
    python generate_cloud_asset.py                         # localhost + defaults
    python generate_cloud_asset.py --ip 192.168.51.40     # remote agent IP
    python generate_cloud_asset.py --ip 192.168.51.40 \\
        --gateway_name DSF-GW01                            # named gateway
    python generate_cloud_asset.py --env-file /path/.env  # explicit env file

Relevant env vars (all have defaults, .env file loaded automatically)
----------------------------------------------------------------------
  FLOCI_DEFAULT_ACCOUNT_ID  AWS account ID used for ARNs     (default: 000000000000)
  FAM_ACCOUNT_ID            Fallback account ID               (default: 000000000000)
  AWS_DEFAULT_REGION        AWS region                        (default: us-east-1)
  AWS_ACCESS_KEY_ID         LocalStack / Floci access key     (default: test)
  AWS_SECRET_ACCESS_KEY     LocalStack / Floci secret key     (default: test)
  FLOCI_HOST_PORT           LocalStack port                   (parsed from AWS_ENDPOINT_URL if unset, default 4566)
  AWS_ENDPOINT_URL          Full endpoint URL                 (default: http://localhost:4566)
  ENV_SUFFIX                Appended to the asset display name (e.g. -env1 → floci-local-aws-env1)
"""

import argparse
import csv
import os
import sys
from pathlib import Path
from urllib.parse import urlparse

SCRIPT_DIR = Path(__file__).resolve().parent

HEADERS = [
    "asset_id",
    "asset_display_name",
    "Server Type",
    "Server IP",
    "Server Host Name",
    "Service Name",
    "asset_source",
    "admin_email",
    "Server Port",
    "auth_mechanism",
    "location",
    "region",
    "access_id",
    "secret_key",
    "jsonar_uid_display_name",
    "credentials_endpoint",
    "service_endpoints.logs",
    "service_endpoints.rds",
    "service_endpoints.s3",
]


# ---------------------------------------------------------------------------
# Env helpers
# ---------------------------------------------------------------------------

def load_env_file(path: Path) -> None:
    """Load KEY=value lines from *path*; already-set vars keep priority."""
    if not path.exists():
        return
    with open(path) as fh:
        for raw in fh:
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            key = key.strip()
            val = val.strip().strip('"').strip("'")
            # same rule as 00-env.sh: skip if already in environment
            if key and key not in os.environ:
                os.environ[key] = val


def _port_from_url(url: str) -> str | None:
    """Return the port string from a URL such as http://host:4567, or None."""
    try:
        p = urlparse(url)
        if p.port:
            return str(p.port)
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# Row builder
# ---------------------------------------------------------------------------

def build_row(
    ip: str,
    gateway_name: str,
    account_id: str,
    region: str,
    access_key: str,
    secret_key: str,
    port: str,
    display_name: str,
) -> dict:
    endpoint = f"http://{ip}:{port}"
    iam_arn = f"arn:aws:iam::{account_id}"
    return {
        "asset_id": iam_arn,
        "asset_display_name": display_name,
        "Server Type": "AWS",
        "Server IP": iam_arn,
        "Server Host Name": ip,
        "Service Name": iam_arn,
        "asset_source": "AWS",
        "admin_email": "dbadmin@company.com",
        "Server Port": port,
        "auth_mechanism": "key",
        "location": region,
        "region": region,
        "access_id": access_key,
        "secret_key": secret_key,
        "jsonar_uid_display_name": gateway_name,
        "credentials_endpoint": endpoint,
        "service_endpoints.logs": endpoint,
        "service_endpoints.rds": endpoint,
        "service_endpoints.s3": endpoint,
    }


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------

def write_csv(rows: list[dict], out: Path) -> None:
    with open(out, "w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"[csv]  Written → {out}")
    print("       openpyxl not found — open the CSV in Excel / LibreOffice and Save As .xlsx")
    print("       Or install it:  pip install openpyxl")


def write_xlsx(rows: list[dict], out: Path) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assets"

    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row.get(h, "") for h in HEADERS])

    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 60)

    wb.save(out)
    print(f"[xlsx] Written → {out}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        prog="generate_cloud_asset.py",
        description=(
            "Generate a DSF / Floci cloud-asset import spreadsheet.\n\n"
            "Env vars are read from the shell environment first, then from a .env\n"
            "file next to this script (same override order as 00-env.sh)."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "examples:\n"
            "  python generate_cloud_asset.py\n"
            "  python generate_cloud_asset.py --ip 192.168.51.40\n"
            "  python generate_cloud_asset.py --ip 192.168.51.40 --gateway_name DSF-GW01\n"
            "  python generate_cloud_asset.py --env-file /other/path/.env\n"
            "  python generate_cloud_asset.py --output /tmp/my_assets.xlsx\n"
        ),
    )
    parser.add_argument(
        "--ip",
        default="localhost",
        metavar="IP",
        help="Server IP / hostname written into endpoint URLs (default: localhost)",
    )
    parser.add_argument(
        "--gateway_name",
        default="<agentless-gateway-user-defined-name>",
        metavar="NAME",
        help=(
            "DSF agentless-gateway display name "
            "(default: <agentless-gateway-user-defined-name>)"
        ),
    )
    parser.add_argument(
        "--env-file",
        default=None,
        metavar="PATH",
        help="Path to a .env file (default: .env next to this script)",
    )
    parser.add_argument(
        "--output",
        default=None,
        metavar="FILE",
        help=(
            "Output file path. Extension determines format: .xlsx or .csv. "
            "Default: assets/assets_spreadsheet_generated.[xlsx|csv]"
        ),
    )
    args = parser.parse_args()

    # ── load env ──────────────────────────────────────────────────────────
    env_file = Path(args.env_file) if args.env_file else SCRIPT_DIR / ".env"
    load_env_file(env_file)
    if env_file.exists():
        print(f"[env]  Loaded {env_file}")
    else:
        print(f"[env]  No .env found at {env_file}, using shell environment / defaults")

    # ── resolve values ────────────────────────────────────────────────────
    account_id = (
        os.environ.get("FLOCI_DEFAULT_ACCOUNT_ID")
        or os.environ.get("FAM_ACCOUNT_ID")
        or "000000000000"
    )
    region     = os.environ.get("AWS_DEFAULT_REGION", "us-east-1")
    access_key = os.environ.get("AWS_ACCESS_KEY_ID", "test")
    secret_key = os.environ.get("AWS_SECRET_ACCESS_KEY", "test")
    env_suffix = os.environ.get("ENV_SUFFIX", "")
    endpoint_url = os.environ.get("AWS_ENDPOINT_URL", "")
    port = (
        os.environ.get("FLOCI_HOST_PORT")
        or _port_from_url(endpoint_url)
        or "4566"
    )
    display_name = f"floci-local-aws{env_suffix}"

    print()
    print("Environment summary")
    print(f"  account_id   : {account_id}")
    print(f"  region       : {region}")
    print(f"  port         : {port}")
    print(f"  ip           : {args.ip}")
    print(f"  gateway_name : {args.gateway_name}")
    print(f"  display_name : {display_name}")
    print(f"  access_key   : {access_key}")
    print(f"  env_suffix   : {env_suffix!r}")
    print()

    rows = [
        build_row(
            ip=args.ip,
            gateway_name=args.gateway_name,
            account_id=account_id,
            region=region,
            access_key=access_key,
            secret_key=secret_key,
            port=port,
            display_name=display_name,
        )
    ]

    # ── determine output format ───────────────────────────────────────────
    try:
        import openpyxl as _  # noqa: F401
        has_xlsx = True
    except ImportError:
        has_xlsx = False

    if args.output:
        out_path = Path(args.output)
        use_xlsx = out_path.suffix.lower() == ".xlsx" and has_xlsx
    else:
        ext = ".xlsx" if has_xlsx else ".csv"
        out_path = SCRIPT_DIR / "assets" / f"assets_spreadsheet_generated{ext}"
        use_xlsx = has_xlsx

    out_path.parent.mkdir(parents=True, exist_ok=True)

    if use_xlsx:
        write_xlsx(rows, out_path)
    else:
        write_csv(rows, out_path)


if __name__ == "__main__":
    main()
